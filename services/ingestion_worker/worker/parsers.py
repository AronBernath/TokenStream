import io
import logging
import re
from typing import Any, Dict, List

import yaml
from bs4 import BeautifulSoup
from bs4.element import Tag
from openpyxl import load_workbook

from common.retrieval_graph import extract_reference_aliases, unique_aliases
from worker.pdf_parser_service import PdfParserService

_STRIP_DUP_SPACES_RE = re.compile(r"\s+")
_MARKDOWN_HEADING_RE = re.compile(r"^(#{1,6})\s+(.+?)\s*$")
_SLUG_RE = re.compile(r"[^a-z0-9]+")
PARSER_VERSION = "generic-core-v1"
logger = logging.getLogger("ingestion-worker.parsers")


def _clean_text_lines(text: str) -> str:
    text = (text or "").replace("\r", "\n")
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return "\n".join([line.strip() for line in text.splitlines() if line.strip()]).strip()


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return _clean_text_lines(str(value))


def _slug(value: str, fallback: str) -> str:
    slug = _SLUG_RE.sub("-", (value or "").lower()).strip("-")
    return slug or fallback


def source_fingerprint(content_hash: str | None) -> str | None:
    content_hash = (content_hash or "").strip()
    if not content_hash:
        return None
    return f"{PARSER_VERSION}:{content_hash}"


def _source_url(raw: dict, src: dict, anchor: str | None = None) -> str | None:
    url = src.get("url") or raw.get("url") or src.get("local_path") or raw.get("local_path")
    if url and anchor:
        return f"{url}#{anchor}"
    return url


def _node_key(prefix: str, *parts: Any) -> str:
    safe = [str(part).strip() for part in parts if str(part).strip()]
    return ":".join([prefix, *safe])


def _dedupe_edges(edges: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for edge in edges:
        src = str(edge.get("src_node_id") or "")
        edge_type = str(edge.get("edge_type") or "")
        dst_id = str(edge.get("dst_node_id") or "")
        dst_alias = str(edge.get("dst_alias") or "")
        if not src or not edge_type or (not dst_id and not dst_alias):
            continue
        key = (src, edge_type, dst_id, dst_alias)
        if key in seen:
            continue
        seen.add(key)
        out.append(edge)
    return out


def _document_aliases(src: dict, title: str, source_url: str | None) -> List[str]:
    aliases = [src.get("id") or "", title, *(src.get("tags") or [])]
    if source_url:
        aliases.append(source_url)
        aliases.extend(extract_reference_aliases(source_url))
    aliases.extend(extract_reference_aliases(f"{src.get('id', '')} {title}"))
    return unique_aliases(aliases)


def _build_document_node(src: dict, title: str, source_url: str | None) -> Dict[str, Any]:
    doc_id = src.get("id") or source_url or "doc"
    return {
        "node_id": _node_key("doc", doc_id),
        "node_type": "document",
        "label": title or doc_id,
        "aliases": _document_aliases(src, title or doc_id, source_url),
        "metadata": {
            "doc_id": doc_id,
            "doc_type": src.get("doc_type"),
            "source_url": source_url,
            "tags": src.get("tags", []),
        },
    }


def _section_aliases(block: Dict[str, Any]) -> List[str]:
    metadata = block.get("metadata") or {}
    aliases: List[str] = []
    aliases.extend(metadata.get("section_aliases") or [])
    aliases.append(block.get("section_id") or "")
    preview = _STRIP_DUP_SPACES_RE.sub(" ", str(block.get("text") or "")[:240]).strip()
    aliases.extend(extract_reference_aliases(preview))
    return unique_aliases(aliases)


def _build_primary_node(block: Dict[str, Any], src: dict, document_node: Dict[str, Any]) -> Dict[str, Any]:
    doc_id = src.get("id") or block.get("doc_id") or "doc"
    section_id = block.get("section_id")
    if section_id:
        section_kind = str((block.get("metadata") or {}).get("section_kind") or "section")
        return {
            "node_id": _node_key(section_kind, doc_id, section_id),
            "node_type": section_kind,
            "label": str(block.get("title") or section_id),
            "aliases": _section_aliases(block),
            "metadata": {
                "doc_id": doc_id,
                "section_id": section_id,
                "source_url": block.get("source_url"),
            },
        }
    return document_node


def _annotate_graph_metadata(blocks: List[Dict[str, Any]], src: dict, corpus: dict) -> List[Dict[str, Any]]:
    if not blocks:
        return blocks

    title = str(blocks[0].get("title") or corpus.get("title") or src.get("id") or "")
    source_url = blocks[0].get("source_url") or src.get("url")
    document_node = _build_document_node(src, title, source_url)
    prev_primary_id: str | None = None

    for block in blocks:
        metadata = dict(block.get("metadata") or {})
        primary_node = _build_primary_node(block, src, document_node)
        edges: List[Dict[str, Any]] = []

        if primary_node["node_id"] != document_node["node_id"]:
            edges.append(
                {
                    "src_node_id": primary_node["node_id"],
                    "edge_type": "part_of",
                    "dst_node_id": document_node["node_id"],
                    "weight": 1.0,
                    "metadata": {},
                }
            )
            edges.append(
                {
                    "src_node_id": document_node["node_id"],
                    "edge_type": "has_section",
                    "dst_node_id": primary_node["node_id"],
                    "weight": 1.0,
                    "metadata": {},
                }
            )

        if prev_primary_id and prev_primary_id != primary_node["node_id"]:
            edges.append(
                {
                    "src_node_id": prev_primary_id,
                    "edge_type": "next",
                    "dst_node_id": primary_node["node_id"],
                    "weight": 0.45,
                    "metadata": {},
                }
            )
            edges.append(
                {
                    "src_node_id": primary_node["node_id"],
                    "edge_type": "prev",
                    "dst_node_id": prev_primary_id,
                    "weight": 0.45,
                    "metadata": {},
                }
            )

        for alias in unique_aliases(metadata.get("graph_ref_aliases") or []):
            edges.append(
                {
                    "src_node_id": primary_node["node_id"],
                    "edge_type": "refers_to",
                    "dst_alias": alias,
                    "weight": 0.8,
                    "metadata": {"resolver": "alias"},
                }
            )

        metadata["graph_document_node"] = document_node
        metadata["graph_primary_node"] = primary_node
        metadata["graph_edges"] = _dedupe_edges(edges)
        metadata["graph_aliases"] = primary_node.get("aliases") or []
        block["metadata"] = metadata
        prev_primary_id = primary_node["node_id"]
    return blocks


def _base_block(
    *,
    raw: dict,
    src: dict,
    corpus: dict,
    text: str,
    title: str,
    section_id: str,
    fmt: str,
    section_kind: str,
    ordinal: int,
    source_anchor: str | None = None,
    extra_metadata: dict | None = None,
) -> Dict[str, Any]:
    metadata = {
        "format": fmt,
        "parser_version": PARSER_VERSION,
        "section_kind": section_kind,
        "ordinal": ordinal,
        "section_aliases": [section_id, title],
        "graph_ref_aliases": extract_reference_aliases(text),
    }
    if source_anchor:
        metadata["source_anchor"] = source_anchor
    if extra_metadata:
        metadata.update(extra_metadata)
    return {
        "title": title,
        "section_id": section_id,
        "text": text,
        "source_url": _source_url(raw, src, source_anchor),
        "language": src.get("language"),
        "doc_id": src.get("id"),
        "doc_type": src.get("doc_type", fmt),
        "tags": src.get("tags", []),
        "metadata": metadata,
    }


def _select_html_main(soup: BeautifulSoup, rule: dict) -> Tag:
    selectors = [rule.get("main_selector"), "main", "[role='main']", "article", "body"]
    for selector in selectors:
        if not selector:
            continue
        node = soup.select_one(selector)
        if isinstance(node, Tag):
            return node
    return soup


def _tag_text(tag: Tag | None, separator: str = "\n") -> str:
    if tag is None:
        return ""
    return _clean_text_lines(tag.get_text(separator, strip=True))


def _html_title(soup: BeautifulSoup, src: dict, corpus: dict) -> str:
    title = _tag_text(soup.select_one("h1")) or _tag_text(soup.select_one("title"))
    return title or src.get("title") or corpus.get("title", "") or src.get("id") or "document"


def _is_nested_content(tag: Tag) -> bool:
    parent = tag.parent
    while isinstance(parent, Tag):
        if parent.name in {"p", "li", "tr", "pre", "blockquote"}:
            return True
        parent = parent.parent
    return False


def _parse_html_docs(soup: BeautifulSoup, raw: dict, src: dict, corpus: dict, rule: dict) -> List[Dict[str, Any]]:
    main = _select_html_main(soup, rule)
    for selector in rule.get("remove_selectors") or ["script", "style", "noscript", "template"]:
        for el in main.select(selector):
            el.decompose()

    doc_title = _html_title(soup, src, corpus)
    current_heading = doc_title
    current_heading_level = 0
    blocks: List[Dict[str, Any]] = []

    for tag in main.find_all(
        ["h1", "h2", "h3", "h4", "h5", "h6", "p", "li", "tr", "pre", "blockquote"], recursive=True
    ):
        if not isinstance(tag, Tag):
            continue
        if tag.name in {"h1", "h2", "h3", "h4", "h5", "h6"}:
            heading = _tag_text(tag, " ")
            if heading:
                current_heading = heading
                current_heading_level = int(tag.name[1])
            continue
        if _is_nested_content(tag):
            continue

        separator = " | " if tag.name == "tr" else "\n"
        body = _tag_text(tag, separator)
        if not body:
            continue
        ordinal = len(blocks)
        anchor = str(tag.get("id") or "").strip() or None
        section_seed = anchor or f"{tag.name}-{ordinal:04d}-{_slug(current_heading, 'section')}"
        title = current_heading or doc_title
        text = body if body == title else f"{title}\n\n{body}".strip()
        blocks.append(
            _base_block(
                raw=raw,
                src=src,
                corpus=corpus,
                text=text,
                title=title,
                section_id=section_seed,
                fmt="html",
                section_kind=tag.name,
                ordinal=ordinal,
                source_anchor=anchor,
                extra_metadata={
                    "heading": current_heading,
                    "heading_level": current_heading_level,
                    "html_tag": tag.name,
                },
            )
        )

    if blocks:
        return blocks

    text = _clean_text_lines(main.get_text("\n"))
    if not text:
        return []
    return [
        _base_block(
            raw=raw,
            src=src,
            corpus=corpus,
            text=text,
            title=doc_title,
            section_id="root",
            fmt="html",
            section_kind="document",
            ordinal=0,
            extra_metadata={"html_tag": "document"},
        )
    ]


def _parse_text_docs(raw: dict, src: dict, corpus: dict, *, fmt: str = "text") -> List[Dict[str, Any]]:
    content = str(raw.get("content", "") or "")
    title = src.get("title") or corpus.get("title", "") or src.get("id") or "document"
    paragraphs = [_clean_text_lines(part) for part in re.split(r"\n\s*\n", content.replace("\r", "\n"))]
    paragraphs = [p for p in paragraphs if p]
    if not paragraphs and content.strip():
        paragraphs = [_clean_text_lines(content)]

    blocks: List[Dict[str, Any]] = []
    for ordinal, paragraph in enumerate(paragraphs):
        blocks.append(
            _base_block(
                raw=raw,
                src=src,
                corpus=corpus,
                text=paragraph,
                title=title,
                section_id=f"paragraph-{ordinal:04d}",
                fmt=fmt,
                section_kind="paragraph",
                ordinal=ordinal,
            )
        )
    return blocks


def _parse_yaml_docs(raw: dict, src: dict, corpus: dict) -> List[Dict[str, Any]]:
    content = str(raw.get("content", "") or "")
    try:
        parsed = yaml.safe_load(content) or {}
    except Exception as exc:
        logger.warning("Failed to parse YAML for %s: %s", src.get("id"), exc)
        parsed = {}

    title = src.get("title") or corpus.get("title", "") or src.get("id") or "document"
    if isinstance(parsed, dict) and isinstance(parsed.get("info"), dict) and parsed["info"].get("title"):
        title = str(parsed["info"]["title"])

    return [
        _base_block(
            raw=raw,
            src=src,
            corpus=corpus,
            text=content,
            title=title,
            section_id="root",
            fmt="yaml",
            section_kind="document",
            ordinal=0,
        )
    ]


def _parse_markdown_docs(raw: dict, src: dict, corpus: dict) -> List[Dict[str, Any]]:
    content = str(raw.get("content", "") or "")
    default_title = src.get("title") or corpus.get("title", "") or src.get("id") or "document"
    lines = content.splitlines()
    blocks: List[Dict[str, Any]] = []
    current_title = default_title
    current_heading_level = 0
    current_lines: List[str] = []

    def flush_block() -> None:
        nonlocal current_lines
        text = "\n".join(current_lines).strip()
        if not text:
            current_lines = []
            return
        ordinal = len(blocks)
        blocks.append(
            _base_block(
                raw=raw,
                src=src,
                corpus=corpus,
                text=text,
                title=current_title,
                section_id=f"section-{ordinal:04d}-{_slug(current_title, 'section')}",
                fmt="markdown",
                section_kind="heading" if current_heading_level else "document",
                ordinal=ordinal,
                extra_metadata={"heading_level": current_heading_level},
            )
        )
        current_lines = []

    for line in lines:
        heading_match = _MARKDOWN_HEADING_RE.match(line)
        if heading_match:
            flush_block()
            current_heading_level = len(heading_match.group(1))
            current_title = heading_match.group(2).strip()
        current_lines.append(line)

    flush_block()
    if blocks:
        return blocks
    return _parse_text_docs(raw, src, corpus, fmt="markdown")


def _parse_xlsx_docs(raw: dict, src: dict, corpus: dict) -> List[Dict[str, Any]]:
    wb = load_workbook(filename=io.BytesIO(raw["content"]), data_only=True)
    title = src.get("title") or corpus.get("title", "") or src.get("id") or "workbook"
    blocks: List[Dict[str, Any]] = []

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [_clean_cell(value) for value in rows[0]]
        has_headers = sum(1 for value in headers if value) >= 2
        data_rows = rows[1:] if has_headers else rows

        for row_index, row in enumerate(data_rows, start=2 if has_headers else 1):
            cells = [_clean_cell(value) for value in row]
            if not any(cells):
                continue
            parts: List[str] = []
            for col_index, cell in enumerate(cells):
                if not cell:
                    continue
                if has_headers and col_index < len(headers) and headers[col_index]:
                    parts.append(f"{headers[col_index]}: {cell}")
                else:
                    parts.append(cell)
            text = "\n".join(parts).strip()
            ordinal = len(blocks)
            sheet_slug = _slug(sheet.title, "sheet")
            blocks.append(
                _base_block(
                    raw=raw,
                    src=src,
                    corpus=corpus,
                    text=text,
                    title=f"{title} / {sheet.title}",
                    section_id=f"{sheet_slug}-row-{row_index}",
                    fmt="xlsx",
                    section_kind="row",
                    ordinal=ordinal,
                    source_anchor=f"sheet={sheet_slug}&row={row_index}",
                    extra_metadata={
                        "sheet": sheet.title,
                        "row": row_index,
                        "headers": headers if has_headers else [],
                    },
                )
            )
    return blocks


def parse_to_blocks(raw: dict, src: dict, corpus: dict, rules: dict | None = None):
    rules = rules or {}
    fmt = str(raw.get("format", "html") or "html").lower()
    content_hash = raw.get("content_hash")

    if fmt == "yaml":
        blocks = _parse_yaml_docs(raw, src, corpus)
    elif fmt in {"markdown", "md"}:
        blocks = _parse_markdown_docs(raw, src, corpus)
    elif fmt in {"text", "txt"}:
        blocks = _parse_text_docs(raw, src, corpus)
    elif fmt == "xlsx":
        blocks = _parse_xlsx_docs(raw, {**src, "local_path": raw.get("local_path")}, corpus)
    elif fmt == "pdf":
        pdf_rules = (rules or {}).get("pdf") or {}
        blocks = PdfParserService().parse_to_blocks(
            raw,
            {**src, "local_path": raw.get("local_path")},
            corpus,
            rules=pdf_rules,
        )
    elif fmt == "html":
        soup = BeautifulSoup(raw["content"], "html.parser")
        html_rules = (rules or {}).get("html") or {}
        blocks = _parse_html_docs(soup, raw, src, corpus, html_rules)
    else:
        raise ValueError(f"Unsupported format: {fmt}")

    if content_hash:
        for block in blocks:
            metadata = block.setdefault("metadata", {})
            metadata["parser_version"] = PARSER_VERSION
            metadata["source_content_hash"] = content_hash
            fingerprint = source_fingerprint(content_hash)
            if fingerprint:
                metadata["source_fingerprint"] = fingerprint

    return _annotate_graph_metadata(blocks, src, corpus)
